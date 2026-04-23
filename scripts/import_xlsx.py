"""
Import account settings and follow targets from thistlegroup-data.xlsx
into the Railway PostgreSQL database.

Usage:
    railway run python scripts/import_xlsx.py

Accounts must already exist in the DB (created via dashboard).
User: ahmartin@gmail.com
"""

import asyncio
import datetime
import json
import os
import pathlib
import uuid

import asyncpg
import openpyxl

# ── config ────────────────────────────────────────────────────────────────────

XLSX_PATH = pathlib.Path(__file__).parent.parent / "thistlegroup-data.xlsx"
USER_EMAIL = "ahmartin@gmail.com"
ACCOUNTS = ["triagespirits", "theburleybar", "marketalleygin", "timeforashifty"]
BATCH_SIZE = 500

# ── data mappings ─────────────────────────────────────────────────────────────

ACTION_TYPE_MAP = {
    "like": "like",
    "follow": "follow",
    "unfollow": "unfollow",
}

ACTION_TARGET_MAP = {
    "post[homepage]": "posts [homepage]",
    "post[topics]": "posts [topics]",
    "suggested": "suggested users",
    "database": "previous follows",
}

STATUS_MAP = {
    "done": "done",
    "private": "skipped",
    "pending": "following",
}


def parse_delay(val) -> tuple[int, int]:
    """Parse delay cell: '90/15' → (90, 15), '60' → (60, 0), datetime → (60, 0)."""
    if val is None:
        return 60, 0
    s = str(val).strip()
    if "/" in s:
        parts = s.split("/")
        try:
            return int(parts[0]), int(parts[1])
        except ValueError:
            return 60, 0
    try:
        return int(float(s)), 0
    except (ValueError, TypeError):
        return 60, 0


def parse_follow_back(val) -> bool | None:
    """Coerce follow_back cell to bool; skip formula strings → None."""
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    s = str(val).strip().lower()
    if s in ("yes", "true", "1"):
        return True
    if s in ("no", "false", "0"):
        return False
    return None  # formula string or unknown


def to_date(val) -> datetime.date | None:
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    return None


def to_time(val) -> datetime.time | None:
    if val is None:
        return None
    if isinstance(val, datetime.time):
        return val
    if isinstance(val, datetime.datetime):
        return val.time()
    return None


def build_actions(row) -> list[dict]:
    """Build 4 action blocks from settings row (indices 15-34)."""
    actions = []
    for i in range(4):
        base = 15 + i * 5
        raw_type = str(row[base + 1] or "").strip()
        raw_target = str(row[base + 2] or "").strip()
        enabled = row[base]
        if isinstance(enabled, str):
            enabled = enabled.strip().upper() in ("TRUE", "1", "YES", "ON")
        else:
            enabled = bool(enabled)

        fixed = row[base + 3]
        var = row[base + 4]

        actions.append({
            "enabled": enabled,
            "type": ACTION_TYPE_MAP.get(raw_type, raw_type),
            "target": ACTION_TARGET_MAP.get(raw_target, raw_target),
            "fixed_count": int(fixed) if fixed is not None else 0,
            "variable_count": int(var) if var is not None else 0,
        })
    return actions


def get_settings_row(ws, account_name: str) -> tuple | None:
    """Find the row for account_name in the settings sheet (rows start at row 3)."""
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] == account_name:
            return row
    return None


# ── main ──────────────────────────────────────────────────────────────────────

async def main():
    # Prefer public URL when running locally (private URL is only reachable inside Railway)
    db_url = os.environ.get("DATABASE_PUBLIC_URL") or os.environ.get("DATABASE_URL")
    if not db_url:
        raise SystemExit("DATABASE_URL not set. Run via: railway run python scripts/import_xlsx.py")

    # asyncpg requires postgresql:// not postgres://
    asyncpg_url = db_url.replace("postgres://", "postgresql://", 1)

    print(f"Connecting to database…")
    conn = await asyncpg.connect(asyncpg_url)

    try:
        # 1. Look up user
        user = await conn.fetchrow(
            "SELECT id FROM users WHERE email = $1", USER_EMAIL
        )
        if not user:
            raise SystemExit(f"User '{USER_EMAIL}' not found in database.")
        user_id = user["id"]
        print(f"Found user: {USER_EMAIL} ({user_id})")

        # 2. Load workbook
        print(f"Loading {XLSX_PATH.name}…")
        wb = openpyxl.load_workbook(str(XLSX_PATH), read_only=True, data_only=True)
        settings_ws = wb["settings"]

        for acct_name in ACCOUNTS:
            print(f"\n-- {acct_name} --")

            # Look up account
            account = await conn.fetchrow(
                "SELECT id FROM accounts WHERE name = $1 AND user_id = $2",
                acct_name, user_id
            )
            if not account:
                print(f"  SKIP: account '{acct_name}' not found in DB (create it in the dashboard first).")
                continue
            account_id = account["id"]
            print(f"  account_id: {account_id}")

            # Parse settings row
            row = get_settings_row(settings_ws, acct_name)
            if not row:
                print(f"  SKIP: no settings row found in xlsx for '{acct_name}'.")
                continue

            schedule_days = str(row[9] or "daily").strip()
            schedule_start = to_time(row[11])
            schedule_end = to_time(row[12])
            delay_base, delay_random = parse_delay(row[13])
            max_runs = int(row[14]) if row[14] is not None else 1
            unfollow_days = int(row[35]) if row[35] is not None else 30
            list_tab = str(row[37]).strip() if row[37] else None
            account_group = str(row[38]).strip() if row[38] else None
            account_list_tab = str(row[39]).strip() if row[39] else None
            topics = str(row[40]).strip() if row[40] else None
            actions = build_actions(row)

            print(f"  schedule: {schedule_days} {schedule_start}–{schedule_end}, delay {delay_base}/{delay_random}, max {max_runs}/day")
            action_summary = [f"{a['type']} ({a['target']})" for a in actions]
            print(f"  actions: {action_summary}")

            # Upsert account_settings
            await conn.execute("""
                INSERT INTO account_settings (
                    id, account_id, user_id,
                    schedule_days, schedule_start, schedule_end,
                    delay_base_minutes, delay_random_minutes, max_runs_per_day,
                    actions, unfollow_days,
                    list_tab, account_group, account_list_tab, topics,
                    updated_at
                ) VALUES (
                    $1, $2, $3,
                    $4, $5, $6,
                    $7, $8, $9,
                    $10, $11,
                    $12, $13, $14, $15,
                    NOW()
                )
                ON CONFLICT (account_id) DO UPDATE SET
                    schedule_days = EXCLUDED.schedule_days,
                    schedule_start = EXCLUDED.schedule_start,
                    schedule_end = EXCLUDED.schedule_end,
                    delay_base_minutes = EXCLUDED.delay_base_minutes,
                    delay_random_minutes = EXCLUDED.delay_random_minutes,
                    max_runs_per_day = EXCLUDED.max_runs_per_day,
                    actions = EXCLUDED.actions,
                    unfollow_days = EXCLUDED.unfollow_days,
                    list_tab = EXCLUDED.list_tab,
                    account_group = EXCLUDED.account_group,
                    account_list_tab = EXCLUDED.account_list_tab,
                    topics = EXCLUDED.topics,
                    updated_at = NOW()
            """,
                uuid.uuid4(), account_id, user_id,
                schedule_days, schedule_start, schedule_end,
                delay_base, delay_random, max_runs,
                json.dumps(actions), unfollow_days,
                list_tab, account_group, account_list_tab, topics,
            )
            print(f"  account_settings upserted.")

            # Import follow_targets
            if acct_name not in wb.sheetnames:
                print(f"  No sheet named '{acct_name}' — skipping follow_targets.")
                continue

            ft_ws = wb[acct_name]
            existing = await conn.fetchval(
                "SELECT COUNT(*) FROM follow_targets WHERE account_id = $1", account_id
            )
            if existing > 0:
                print(f"  follow_targets: {existing} rows already exist — skipping import.")
                continue

            batch = []
            total = 0
            skipped = 0

            for row in ft_ws.iter_rows(min_row=3, values_only=True):
                handle = row[0]
                if not handle or not isinstance(handle, str):
                    continue

                source = str(row[1]).strip() if row[1] else None
                raw_status = str(row[2] or "pending").strip().lower()
                status = STATUS_MAP.get(raw_status, "pending")
                follow_date = to_date(row[3])
                unfollow_date = to_date(row[4])
                follow_back = parse_follow_back(row[5])

                if follow_back is None and row[5] is not None:
                    # formula string — treat as unknown
                    follow_back = None

                batch.append((
                    uuid.uuid4(), user_id, account_id,
                    handle, source, status,
                    follow_date, unfollow_date, follow_back,
                ))

                if len(batch) >= BATCH_SIZE:
                    await conn.executemany("""
                        INSERT INTO follow_targets
                            (id, user_id, account_id, target_handle, source, status,
                             follow_date, unfollow_date, follow_back)
                        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9)
                    """, batch)
                    total += len(batch)
                    print(f"  …{total} rows inserted")
                    batch = []

            if batch:
                await conn.executemany("""
                    INSERT INTO follow_targets
                        (id, user_id, account_id, target_handle, source, status,
                         follow_date, unfollow_date, follow_back)
                    VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9)
                """, batch)
                total += len(batch)

            print(f"  follow_targets: {total} rows inserted, {skipped} skipped.")

        # ── Import runlog ────────────────────────────────────────────────────
        print("\n-- importing runlog --")
        if "runlog" not in wb.sheetnames:
            print("  No 'runlog' sheet found — skipping.")
        else:
            runlog_ws = wb["runlog"]

            # Build account name → id map (all accounts for user)
            account_map = {}
            all_accounts = await conn.fetch(
                "SELECT id, name FROM accounts WHERE user_id = $1", user_id
            )
            for acct in all_accounts:
                account_map[acct["name"]] = acct["id"]
            print(f"  found {len(account_map)} accounts: {list(account_map.keys())}")

            batch = []
            total = 0
            skipped = 0

            for row in runlog_ws.iter_rows(min_row=3, values_only=True):
                run_date = to_date(row[0])
                if not run_date:
                    continue

                acct_name = str(row[4] or "").strip()
                if acct_name not in account_map:
                    skipped += 1
                    continue
                account_id = account_map[acct_name]

                run_seq = int(row[1]) if row[1] is not None else 1
                start_t = to_time(row[2])
                end_t = to_time(row[3])

                # Combine date + time into datetime
                start_dt = datetime.datetime.combine(run_date, start_t) if start_t else None
                end_dt = datetime.datetime.combine(run_date, end_t) if end_t else None

                def safe_int(v):
                    if v is None:
                        return 0
                    try:
                        return int(float(v))
                    except (ValueError, TypeError):
                        return 0

                a1_type = str(row[5]).strip() if row[5] else None
                a1_count = safe_int(row[6])
                a2_type = str(row[7]).strip() if row[7] else None
                a2_count = safe_int(row[8])
                a3_type = str(row[9]).strip() if row[9] else None
                a3_count = safe_int(row[10])
                a4_type = str(row[11]).strip() if row[11] else None
                a4_count = safe_int(row[12])

                error_msg = str(row[16]).strip() if row[16] else None

                batch.append((
                    uuid.uuid4(), user_id, account_id,
                    run_date, run_seq, start_dt, end_dt,
                    a1_type, a1_count, a2_type, a2_count,
                    a3_type, a3_count, a4_type, a4_count,
                    error_msg,
                ))

                if len(batch) >= BATCH_SIZE:
                    await conn.executemany("""
                        INSERT INTO session_logs
                            (id, user_id, account_id,
                             run_date, run_sequence, start_time, end_time,
                             action_1_type, action_1_count, action_2_type, action_2_count,
                             action_3_type, action_3_count, action_4_type, action_4_count,
                             error_message)
                        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16)
                    """, batch)
                    total += len(batch)
                    print(f"  …{total} rows inserted")
                    batch = []

            if batch:
                await conn.executemany("""
                    INSERT INTO session_logs
                        (id, user_id, account_id,
                         run_date, run_sequence, start_time, end_time,
                         action_1_type, action_1_count, action_2_type, action_2_count,
                         action_3_type, action_3_count, action_4_type, action_4_count,
                         error_message)
                    VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16)
                """, batch)
                total += len(batch)

            print(f"  session_logs: {total} rows inserted, {skipped} skipped.")

        print("\nDone.")

    finally:
        await conn.close()


if __name__ == "__main__":
    asyncio.run(main())
